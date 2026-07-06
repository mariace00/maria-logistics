#!/usr/bin/env python3
"""
derive_shipment_master.py

Builds the shipment master list DIRECTLY from the three real SAP source sheets —
no dependency on Final or PROJECT_TRACKING, which are just manually-maintained
copies of this same information.

  ZSD29  = the open-shipments list. One row per OT (SAP "Transport"): weight,
           destination, planned ETA, freight forwarder, and the SO numbers
           riding on that transport (Ped_1..Ped_19).

  ZMM48  = material-in-transit detail. One row per part per OT. This is the
           AUTHORITATIVE source for container ID ("Container identifier") and
           transit date — it's never truncated, unlike ZSD28's notes field.

  ZSD28  = one row per Sales Order: customer, project code, contact, and a
           "Logistic Notes" field. On many rows, someone has pasted a tab-
           separated freight update into that field in the same shape as the
           old Final sheet's columns: OT <TAB> container/tracking# <TAB> dest
           <TAB> POD <TAB> mode <TAB> carrier. IMPORTANT: this SAP field caps
           out around 50-60 characters, so carrier names get cut off
           ("Hapag-" instead of "Hapag-Lloyd") and long notes get truncated.
           This script recovers what it can and flags the rest as "raw_note"
           instead of guessing.

This script produces the shipment MASTER data only — OT, container, SO/customer/
project linkage, weight, destination, planned ETA. It deliberately leaves
latest_log / actual-eta / pod / vessel blank for ocean shipments, because those
fields aren't in any of these three sheets — they only exist by actually
checking the carrier's tracking page. That's what api/cron/poll.js (hapag.js /
msc.js) is for. Together: this script answers "what do we need to track", and
the cron job answers "what's its status right now".
"""
import sys
import re
import json
import datetime
import openpyxl
from collections import defaultdict

CONTAINER_RE = re.compile(r'^[A-Z]{4}\d{6,7}$')

KNOWN_CARRIER_FRAGMENTS = {
    "hapag": "Hapag-Lloyd",
    "cma cgm": "CMA CGM",
    "msc": "MSC",
    "maersk": "Maersk",
    "oocl": "OOCL",
}

AWB_PREFIXES = {
    "020": "Lufthansa Cargo", "014": "Air Canada Cargo", "176": "Emirates SkyCargo",
    "125": "British Airways World Cargo", "001": "American Airlines Cargo",
    "006": "Delta Cargo", "016": "United Cargo", "074": "KLM Cargo", "057": "Air France Cargo",
}


def d(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    return v


def complete_carrier(fragment):
    """Recover a full carrier name from a truncated ZSD28 note fragment."""
    if not fragment:
        return None
    frag_lower = fragment.lower()
    for key, full in KNOWN_CARRIER_FRAGMENTS.items():
        if frag_lower.startswith(key):
            return full
    return fragment or None


def detect_air_carrier(tracking):
    if not tracking or str(tracking).strip().lower() in ("tbd", ""):
        return None, None
    t = str(tracking).strip()
    compact = t.replace(" ", "")
    if compact.upper().startswith("1Z"):
        return "UPS", f"https://www.ups.com/track?loc=en_US&tracknum={compact}"
    if re.match(r'^\d{12}$', compact):
        return "FedEx", f"https://www.fedex.com/fedextrack/?trknbr={compact}"
    m = re.match(r'^(\d{3})-?(\d{7,8})$', t)
    if m:
        prefix, serial = m.group(1), m.group(2)
        carrier = AWB_PREFIXES.get(prefix, f"Air carrier (AWB prefix {prefix})")
        if prefix == "020":
            url = f"https://www.lufthansa-cargo.com/en/online-services/track-and-trace?awb={prefix}-{serial}"
        elif prefix == "014":
            url = f"https://www.aircanada.com/cargo/tracking/en/awb/{prefix}-{serial}"
        else:
            url = f"https://www.track-trace.com/aircargo?data={prefix}-{serial}"
        return carrier, url
    return None, None


def parse_logistic_note(note):
    """
    Attempts to pull a structured freight update out of a ZSD28 Logistic Notes
    cell. Returns None if the note is just a free-text comment (e.g. "not
    released yet") rather than a pasted shipment update.
    """
    if not note or '\t' not in note:
        return None
    parts = note.split('\t')
    first = parts[0].strip()
    if not re.match(r'^\d{9,10}$', first):
        return None

    ot = int(first)
    tracking_raw = parts[1].strip() if len(parts) > 1 else None
    rest = [p.strip() for p in parts[2:] if p.strip()]

    # tracking sometimes has a trailing date glued on with no tab, e.g.
    # "8733 6358 7326 6/29/2026"
    date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})$', tracking_raw or '')
    note_date = None
    tracking = tracking_raw
    if date_match:
        note_date = date_match.group(1)
        tracking = tracking_raw[:date_match.start()].strip()

    dest = rest[0] if rest else None
    pod = rest[1] if len(rest) > 1 else None
    mode = rest[2] if len(rest) > 2 else None
    carrier_frag = rest[3] if len(rest) > 3 else None

    # Sometimes trailing date shows up as its own fragment instead
    for r in rest:
        m = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})$', r)
        if m:
            note_date = r

    return {
        "ot": ot, "tracking": tracking, "dest": dest, "pod": pod,
        "mode": mode, "carrier_fragment": carrier_frag, "note_date": note_date,
        "raw": note,
    }


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---------- ZSD28: SO-level data + parsed logistic notes ----------
    zsd28 = wb['ZSD28']
    z28_rows = list(zsd28.iter_rows(min_row=2, max_row=zsd28.max_row, values_only=True))
    z28_header = [c.value for c in zsd28[1]]
    idx = {name: z28_header.index(name) for name in
           ['Logistic Notes', 'Contact name', 'Ship-to party name', 'Ship-to party city', 'Project']}

    so_master = {}          # so_number -> {customer, project, contact, city, weight, raw_note}
    ot_from_notes = defaultdict(list)   # ot -> [ parsed_note_dict + so context ]

    for r in z28_rows:
        so = r[0]
        if so is None or str(so).strip() == '':
            continue
        try:
            so = int(so)
        except (ValueError, TypeError):
            continue
        customer, weight, notes = r[1], r[4], r[idx['Logistic Notes']]
        contact, ship_to, city = r[idx['Contact name']], r[idx['Ship-to party name']], r[idx['Ship-to party city']]
        project = r[idx['Project']]

        so_master[so] = {
            "customer": customer, "project": project, "contact": contact,
            "city": city, "weight": weight, "raw_note": notes if isinstance(notes, str) else None,
        }

        parsed = parse_logistic_note(notes) if isinstance(notes, str) else None
        if parsed:
            parsed["so"] = so
            ot_from_notes[parsed["ot"]].append(parsed)

    # ---------- ZSD29: OT-level open-shipment master ----------
    zsd29 = wb['ZSD29']
    z29_rows = list(zsd29.iter_rows(min_row=2, max_row=zsd29.max_row, values_only=True))
    ot_header = {}
    for r in z29_rows:
        ot = int(r[0])
        peds = [int(p) for p in r[16:35] if p is not None]
        ot_header[ot] = {
            "weight": r[5], "agency": r[6], "dest": r[8], "state_desc": r[11],
            "customer": r[12], "eta": d(r[3]), "peds": peds,
        }

    # ---------- ZMM48: authoritative OT -> container mapping ----------
    zmm48 = wb['ZMM48']
    z48_rows = list(zmm48.iter_rows(min_row=2, max_row=zmm48.max_row, values_only=True))
    z48_header = [c.value for c in zmm48[1]]
    si, ci, ti = z48_header.index('Shipment Number'), z48_header.index('Container identifier'), z48_header.index('Transit date')

    ot_container = {}
    for r in z48_rows:
        ot_raw, cont, tdate = r[si], r[ci], r[ti]
        if not ot_raw:
            continue
        ot = int(ot_raw)
        cont_clean = cont.replace(' ', '').strip() if isinstance(cont, str) else cont
        flag = bool(cont_clean and not CONTAINER_RE.match(cont_clean))
        if cont_clean and not ot_container.get(ot, {}).get('container'):
            ot_container[ot] = {"container": cont_clean, "transit_date": d(tdate), "container_flagged": flag}
        elif ot not in ot_container:
            ot_container[ot] = {"container": None, "transit_date": d(tdate), "container_flagged": False}

    # ---------- Merge into shipment master ----------
    all_ots = set(ot_header.keys()) | set(ot_from_notes.keys()) | set(ot_container.keys())
    shipments = {}

    for ot in all_ots:
        header = ot_header.get(ot)
        notes = ot_from_notes.get(ot, [])
        zmm = ot_container.get(ot)

        container = (zmm or {}).get("container")
        if not container:
            for n in notes:
                if n["tracking"] and CONTAINER_RE.match(n["tracking"]):
                    container = n["tracking"]
                    break

        ship_type = "Ocean" if container else "Air"

        tracking = container
        if ship_type == "Air":
            tracking = notes[0]["tracking"] if notes else None

        dest = (header or {}).get("dest") or (notes[0]["dest"] if notes and notes[0]["dest"] else None)
        weight = (header or {}).get("weight")
        eta = (header or {}).get("eta") or (notes[0]["note_date"] if notes and notes[0]["note_date"] else None)
        pod = notes[0]["pod"] if notes and notes[0].get("pod") else None
        mode = notes[0]["mode"] if notes and notes[0].get("mode") else None

        carrier = None
        air_url = None
        if ship_type == "Ocean":
            frag = next((n["carrier_fragment"] for n in notes if n.get("carrier_fragment")), None)
            carrier = complete_carrier(frag)
        else:
            carrier, air_url = detect_air_carrier(tracking)

        # linked SOs: prefer ZSD29's Ped_x list (complete), fall back to notes' SO
        so_numbers = (header or {}).get("peds") or [n["so"] for n in notes]
        orders = []
        seen = set()
        for so in so_numbers:
            if so in seen:
                continue
            seen.add(so)
            info = so_master.get(so, {})
            orders.append({
                "so": so, "customer": info.get("customer"), "project": info.get("project"),
                "contact": info.get("contact"), "city": info.get("city"),
                "raw_note": info.get("raw_note") if not (header or notes) else None,
            })

        shipments[ot] = {
            "ot": ot, "container": container, "tracking": tracking, "shipType": ship_type,
            "carrier": carrier, "airTrackingUrl": air_url, "dest": dest, "weight": weight,
            "plannedEta": eta, "pod": pod, "mode": mode, "orders": orders,
            "containerFlagged": (zmm or {}).get("container_flagged", False),
            "_sources": {
                "in_zsd29": ot in ot_header, "in_zmm48": ot in ot_container,
                "in_zsd28_notes": ot in ot_from_notes,
            },
        }

    # ---------- Secondary pass: recover OTs only mentioned in free-text (non-tab) notes ----------
    # e.g. "Shipment 2960081867 planned, not departed yet" — no structured tracking data,
    # but still tells us a real shipment exists and which SO/customer it belongs to.
    mention_re = re.compile(r'\b(11\d{8}|29\d{8})\b')
    recovered_from_mentions = []
    for r in z28_rows:
        so_raw, notes = r[0], r[idx['Logistic Notes']]
        if not isinstance(notes, str) or '\t' in notes:
            continue  # already handled by the structured parser above
        for m in mention_re.finditer(notes):
            ot = int(m.group(1))
            if ot in shipments:
                continue
            try:
                so = int(so_raw)
            except (ValueError, TypeError):
                continue
            info = so_master.get(so, {})
            shipments[ot] = {
                "ot": ot, "container": None, "tracking": None, "shipType": "Unknown",
                "carrier": None, "airTrackingUrl": None, "dest": None, "weight": info.get("weight"),
                "plannedEta": None, "pod": None, "mode": None,
                "orders": [{"so": so, "customer": info.get("customer"), "project": info.get("project"),
                            "contact": info.get("contact"), "city": info.get("city"), "raw_note": notes}],
                "containerFlagged": False,
                "_sources": {"in_zsd29": False, "in_zmm48": False, "in_zsd28_notes": False, "mention_only": True},
            }
            recovered_from_mentions.append(ot)

    return shipments, so_master, recovered_from_mentions


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python3 derive_shipment_master.py <path-to-xlsm>")
        sys.exit(1)
    shipments, so_master, recovered = main(sys.argv[1])
    print(f"Derived {len(shipments)} shipments directly from ZSD28 + ZSD29 + ZMM48.")
    print(f"  {len([s for s in shipments.values() if s['_sources']['in_zsd29']])} found in ZSD29 (open transports)")
    notes_only = [ot for ot, s in shipments.items() if s['_sources']['in_zsd28_notes'] and not s['_sources']['in_zsd29']]
    print(f"  {len(notes_only)} recovered from structured ZSD28 notes only (not yet a formal Transport): {notes_only}")
    print(f"  {len(recovered)} recovered from free-text OT mentions (no tracking data available): {recovered}")
    flagged = [s['ot'] for s in shipments.values() if s.get('containerFlagged')]
    if flagged:
        print(f"  ⚠ {len(flagged)} container ID(s) look malformed in the source data, needs manual check: {flagged}")
